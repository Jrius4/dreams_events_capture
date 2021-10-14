from openpyxl import load_workbook, Workbook

wb_tracker = load_workbook("all_girls.xlsx")
wb_events = load_workbook("eventList.xlsx")
# eventList.xlsx
# ws_events = wb_tracker.active
# ws_girls = wb_events.active

ws_girls = wb_tracker.active
ws_events = wb_events.active



events_names = []
tracker_names = []

max_rows_e = ws_events.max_row


for row in ws_events.iter_rows(min_row=2,min_col=11,max_col=11,max_row=max_rows_e,values_only=True):
    events_names.append(row[0])

print(events_names)
# print("\n\n")

max_rows_t = ws_girls.max_row
max_colum_t= ws_girls.max_column
mer_in = []
mer_not = []
ws_girls_data = []
for row in ws_girls.iter_rows(min_row=2,min_col=1,max_col=max_colum_t,max_row=1128,values_only=True):
    ws_girls_data.append(row)
    print(row[5])
    if row[5] in events_names:
        # print("yes")
        mer_in.append(row)
    else:
        # print("no")
        mer_not.append(row)


print(mer_in)

print("\n\n")
print("\n\n")
print("\n\n")

print(mer_not)

print(len(mer_not))
print(len(mer_in))

print("\n\n")
print("\n\n")
print("\n\n")
wb_data = Workbook()
# if 'not_in_capture' in wb_tracker.sheetnames:
#     print('not_in_capture exists')
#     ws2 = wb_data['not_in_capture']
# else:
#    ws2 = wb_data.create_sheet(title='not_in_capture')
ws2 = wb_data.active

ws2.cell(row=1,column=1,value='dreams_id')
ws2.cell(row=1,column=2,value='fullnames')
for i,j in enumerate(mer_not):
    # ws2.cell(row=(i+2),column=1,value=j)
    print(f"i:{i},j:{j}")
    for x,y in enumerate(j):
        ws2.cell(row=(i+2),column=(1+x),value=y)
        print(f"i:{i},j:{j},x:{x},y:{y}")
    print("\n\n")
    print("\n\n")
    print("\n\n")

wb_data.save("output/results.xlsx")